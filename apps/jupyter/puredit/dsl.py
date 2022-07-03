from typing import Callable, Iterable, Sequence, TypeVar
from enum import Enum
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
import openpyxl


AggregationMethod = Enum("AggregationMethod", "minimal maximal")

aggregation_method_to_fn: dict[AggregationMethod, Callable] = {
    AggregationMethod.minimal: min,
    AggregationMethod.maximal: max,
}

T = TypeVar("T")


def unzip(seq: Iterable[Sequence[T]]) -> Sequence[Sequence[T]]:
    return list(zip(*seq))


def merge_dictionaries(*args: dict) -> dict:
    result = {}
    for arg in args:
        result.update(arg)
    return result


class Sheet(object):
    def __init__(self, inner: Worksheet):
        self.inner = inner

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        return None

    def take(
        self, sheet_range: str, filter_expression: str
    ) -> Sequence[Sequence[Cell]]:
        filter = compile(filter_expression, filename="<string>", mode="eval")
        return unzip(
            row
            for row in self.inner[sheet_range]
            if eval(filter, {cell.column_letter: cell.value for cell in row})
        )

    def join(
        self,
        sheet_range: str,
        join_on: Sequence[Cell],
        key_expression: str,
        aggregation_method: AggregationMethod,
    ) -> Sequence[Sequence[Cell]]:
        aggregation_fn = aggregation_method_to_fn[aggregation_method]
        key = compile(key_expression, filename="<string>", mode="eval")
        return unzip(
            aggregation_fn(
                (row for row in self.inner[sheet_range]),
                key=lambda row: eval(
                    key,
                    merge_dictionaries(
                        {cell.column_letter: cell.value for cell in row},
                        {join_cell.column_letter: join_cell.value},
                    ),
                ),
            )
            for join_cell in join_on
        )


def load_sheet(file_name: str, sheet_name: str) -> Sheet:
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook[sheet_name]
    return Sheet(sheet)


def create_workbook(sheet_name: str, columns: Sequence[Sequence[Cell]]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.create_sheet(sheet_name, 0)
    values = [[cell.value for cell in column] for column in columns]
    rows = zip(*values)
    for row in rows:
        sheet.append(row)
    return workbook


def store_sheet(file_name: str, sheet_name: str, columns: Sequence[Sequence[Cell]]):
    workbook = create_workbook(sheet_name, columns)
    workbook.save(file_name)


def display(columns: Sequence[Sequence[Cell]]):
    from pandas import DataFrame
    from IPython.display import display

    df = DataFrame(
        {column[0].column_letter: [cell.value for cell in column] for column in columns}
    )
    display(df)
